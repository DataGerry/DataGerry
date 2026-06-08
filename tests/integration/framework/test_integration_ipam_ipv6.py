# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2026 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Integration tests for the IPv6 IPAM read + enforcement paths against a real MongoDB

The unit tests for the validators and overview builders mock the DB; these tests pin the parts
that only run against Mongo for an IPv6 topology: the find_objects query shapes (the subnet
$elemMatch on 'fields', the interface $elemMatch on the dg-ipam-interface MDS rows), the
assigned-IP roll-up, the supernet/subnet overview payloads (IPv6 family, null percentages,
assigned-only IP table), subnet containment/family validation, and the canonicalize-on-store
behaviour of enforce_object_invariants.

A SUPERNET + SUBNET + carrier CmdbType and one IPv6 supernet / subnet / interface object are
seeded directly into the collections; the framework helpers then run against that real data.
"""
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook

from cmdb.database import MongoDatabaseManager
from cmdb.manager import ObjectsManager, TypesManager
from cmdb.models.object_model import (
    CmdbObject,
    CmdbObjectKey,
    CmdbObjectFieldKey,
    CmdbObjectMdsKey,
    CmdbObjectMdsRowKey,
    extract_field_value,
)
from cmdb.models.type_model import CmdbType
from cmdb.models.special_type_model.ipam_constants import (
    SupernetField,
    SubnetField,
    InterfaceField,
    IpamSection,
    IpamOverviewKey,
    IpamBucketLabel,
    IpAddressFamily,
    IpamRowStatus,
    IpamUnassignKey,
    IpamUnassignMode,
    IpamSubnetIpsExport,
)
from cmdb.framework.ipam.subnet_validator import validate_subnet
from cmdb.framework.ipam.supernet_overview import build_supernet_overview
from cmdb.framework.ipam.subnet_overview import build_subnet_overview, build_subnet_sector_ips
from cmdb.framework.ipam.subnet_unassign import unassign_ips_from_subnet
from cmdb.framework.ipam.subnet_export import build_subnet_ips_xlsx
from cmdb.framework.ipam.enforcement import enforce_object_invariants
from cmdb.utils import ValidationErrorKey
# -------------------------------------------------------------------------------------------------------------------- #

SUPERNET_TYPE_ID: int = 9410
SUBNET_TYPE_ID: int = 9411
CARRIER_TYPE_ID: int = 9412

SUPERNET_ID: int = 9420
SUBNET_ID: int = 9421
CARRIER_ID: int = 9422
CANON_SUBNET_ID: int = 9423
UNASSIGN_REF_OWNER_ID: int = 9430
UNASSIGN_ROW_OWNER_ID: int = 9431

SUPERNET_RANGE_V6: str = '2001:db8::/48'
SUBNET_RANGE_V6: str = '2001:db8:0:1::/64'
ASSIGNED_IP_V6: str = '2001:db8:0:1::5'

TYPE_IDS: list[int] = [SUPERNET_TYPE_ID, SUBNET_TYPE_ID, CARRIER_TYPE_ID]
OBJECT_IDS: list[int] = [
    SUPERNET_ID, SUBNET_ID, CARRIER_ID, CANON_SUBNET_ID,
    UNASSIGN_REF_OWNER_ID, UNASSIGN_ROW_OWNER_ID,
]


def _type_doc(public_id: int, name: str, label: str, special_type: str | None) -> dict[str, Any]:
    """Builds a minimal active CmdbType doc; special_type marks it as an IPAM SpecialType."""
    doc: dict[str, Any] = {
        'public_id': public_id,
        'name': name,
        'label': label,
        'author_id': 1,
        'creation_time': datetime.now(timezone.utc),
        'active': True,
        'fields': [{'type': 'text', 'name': 'dg-name', 'label': 'Name'}],
        'render_meta': {'icon': 'fa-cube', 'sections': [], 'summary': {'fields': ['dg-name']}},
        'acl': {'activated': False, 'groups': {'includes': None}},
        'version': '1.0.0',
    }

    if special_type is not None:
        doc['special_type'] = special_type

    return doc


def _field(name: str, value: Any) -> dict[str, Any]:
    """Builds one CmdbObject 'fields' entry."""
    return {'type': 'text', 'name': name, 'value': value}


def _object_doc(public_id: int, type_id: int, fields: list[dict[str, Any]],
                mds: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    """Builds a CmdbObject doc to seed directly into the collection."""
    doc: dict[str, Any] = {
        'public_id': public_id,
        'type_id': type_id,
        'active': True,
        'author_id': 1,
        'creation_time': datetime.now(timezone.utc),
        'version': '1.0.0',
        'fields': fields,
    }

    if mds is not None:
        doc['multi_data_sections'] = mds

    return doc


def _interface_carrier(owner_id: int, ip: str) -> dict[str, Any]:
    """Builds a carrier object with one dg-ipam-interface row referencing the test subnet at ``ip``."""
    return _object_doc(owner_id, CARRIER_TYPE_ID, [_field('dg-name', f'host-{owner_id}')], mds=[{
        CmdbObjectMdsKey.SECTION_ID: IpamSection.INTERFACE,
        CmdbObjectMdsKey.VALUES: [{CmdbObjectMdsRowKey.DATA: [
            _field(InterfaceField.SUBNET, SUBNET_ID),
            _field(InterfaceField.IP, ip),
        ]}],
    }])


@pytest.fixture(scope='module', autouse=True)
def _seed_ipv6_topology(database_manager: MongoDatabaseManager, database_name: str):
    """Seeds the IPAM types + an IPv6 supernet / subnet / interface object, cleaning up after."""
    types = database_manager.get_collection(CmdbType.COLLECTION, database_name)
    objects = database_manager.get_collection(CmdbObject.COLLECTION, database_name)

    types.insert_many([
        _type_doc(SUPERNET_TYPE_ID, 'it-supernet', 'Supernet', 'SUPERNET'),
        _type_doc(SUBNET_TYPE_ID, 'it-subnet', 'Subnet', 'SUBNET'),
        _type_doc(CARRIER_TYPE_ID, 'it-carrier', 'Carrier', None),
    ])

    objects.insert_many([
        _object_doc(SUPERNET_ID, SUPERNET_TYPE_ID, [
            _field(SupernetField.NAME, 'sn6'),
            _field(SupernetField.TYPE, IpAddressFamily.IPV6),
            _field(SupernetField.NETWORK_RANGE, SUPERNET_RANGE_V6),
        ]),
        _object_doc(SUBNET_ID, SUBNET_TYPE_ID, [
            _field(SubnetField.NAME, 'sub6'),
            _field(SubnetField.TYPE, IpAddressFamily.IPV6),
            _field(SubnetField.PARENT_SUPERNET, SUPERNET_ID),
            _field(SubnetField.NETWORK_RANGE, SUBNET_RANGE_V6),
        ]),
        _object_doc(CARRIER_ID, CARRIER_TYPE_ID, [_field('dg-name', 'host6')], mds=[{
            'section_id': IpamSection.INTERFACE,
            'values': [{'data': [
                _field(InterfaceField.SUBNET, SUBNET_ID),
                _field(InterfaceField.IP, ASSIGNED_IP_V6),
            ]}],
        }]),
    ])

    yield

    types.delete_many({'public_id': {'$in': TYPE_IDS}})
    objects.delete_many({'public_id': {'$in': OBJECT_IDS}})


@pytest.fixture(name='objects_manager')
def fixture_objects_manager(database_manager: MongoDatabaseManager) -> ObjectsManager:
    """Provides an ObjectsManager wired to the test database."""
    return ObjectsManager(database_manager)


@pytest.fixture(name='types_manager')
def fixture_types_manager(database_manager: MongoDatabaseManager) -> TypesManager:
    """Provides a TypesManager wired to the test database."""
    return TypesManager(database_manager)


# -------------------------------------------------------------------------------------------------------------------- #
#                                            SUPERNET OVERVIEW (IPv6)                                                  #
# -------------------------------------------------------------------------------------------------------------------- #
def test_supernet_overview_ipv6_reports_family_null_percentages_and_finds_subnet(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The real find_objects queries surface the IPv6 subnet; KPI strip is family ipv6 with null percentages"""
    payload = build_supernet_overview(objects_manager, types_manager, SUPERNET_ID)

    summary = payload[IpamOverviewKey.SUPERNET]
    assert summary[IpamOverviewKey.SUBNET_TYPE] == IpAddressFamily.IPV6
    assert summary[IpamOverviewKey.USED_PERCENT] is None
    assert summary[IpamOverviewKey.FREE_PERCENT] is None
    assert summary[IpamOverviewKey.UTILIZATION_PERCENT] is None

    rows = payload[IpamOverviewKey.SUBNETS][IpamOverviewKey.ROWS]
    subnet_row = next(r for r in rows if r[CmdbObjectKey.PUBLIC_ID] == SUBNET_ID)
    assert subnet_row[IpamOverviewKey.SUBNET_TYPE] == IpAddressFamily.IPV6
    assert subnet_row[IpamOverviewKey.USAGE_PERCENT] is None
    # the seeded interface row references this subnet, so its used-IP roll-up is 1
    assert subnet_row[IpamOverviewKey.USED_IPS] == 1


# -------------------------------------------------------------------------------------------------------------------- #
#                                             SUBNET OVERVIEW (IPv6)                                                   #
# -------------------------------------------------------------------------------------------------------------------- #
def test_subnet_overview_ipv6_is_assigned_only_with_null_percentages(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The IPv6 subnet IP table lists only the assigned address; type distribution drops Free + nulls %"""
    payload = build_subnet_overview(objects_manager, types_manager, SUBNET_ID)

    subnet = payload[IpamOverviewKey.SUBNET]
    assert subnet[IpamOverviewKey.SUBNET_TYPE] == IpAddressFamily.IPV6
    assert subnet[IpamOverviewKey.USED_IPS] == 1

    ips_block = payload[IpamOverviewKey.IPS]
    assert ips_block[IpamOverviewKey.TOTAL] == 1
    assert [r[IpamOverviewKey.IP] for r in ips_block[IpamOverviewKey.ROWS]] == [ASSIGNED_IP_V6]

    distribution = payload[IpamOverviewKey.TYPE_DISTRIBUTION]
    labels = [b[IpamOverviewKey.LABEL] for b in distribution]
    assert IpamBucketLabel.FREE not in labels
    assert all(b[IpamOverviewKey.PERCENTAGE] is None for b in distribution)

    grid = payload[IpamOverviewKey.IP_DISTRIBUTION]
    assert grid[IpamOverviewKey.RANGES][0][IpamOverviewKey.IP_START] == '2001:db8:0:1::'
    assert all(
        sector[IpamOverviewKey.PERCENTAGE] is None
        for range_block in grid[IpamOverviewKey.RANGES]
        for sector in range_block[IpamOverviewKey.SECTORS]
    )


# -------------------------------------------------------------------------------------------------------------------- #
#                                          SECTOR DRILL-DOWN (IPv6)                                                    #
# -------------------------------------------------------------------------------------------------------------------- #
def test_subnet_sector_ipv6_returns_only_assigned_ips_in_the_window(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """Drilling into the first sector of the IPv6 /64 returns only its assigned address, from real Mongo"""
    payload = build_subnet_sector_ips(objects_manager, types_manager, SUBNET_ID, '2001:db8:0:1::')

    assert payload[IpamOverviewKey.SECTOR][IpamOverviewKey.IP_START] == '2001:db8:0:1::'
    ips_block = payload[IpamOverviewKey.IPS]
    assert ips_block[IpamOverviewKey.TOTAL] == 1
    assert [r[IpamOverviewKey.IP] for r in ips_block[IpamOverviewKey.ROWS]] == [ASSIGNED_IP_V6]


# -------------------------------------------------------------------------------------------------------------------- #
#                                            SUBNET VALIDATION (IPv6)                                                  #
# -------------------------------------------------------------------------------------------------------------------- #
def test_validate_subnet_ipv6_candidate_inside_ipv6_supernet_is_valid(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """A non-overlapping IPv6 candidate strictly inside the IPv6 supernet validates clean"""
    errors = validate_subnet(
        objects_manager, types_manager, '2001:db8:0:2::/64',
        parent_supernet_id=SUPERNET_ID, subnet_type=IpAddressFamily.IPV6,
    )

    assert errors == []


def test_validate_subnet_ipv4_candidate_under_ipv6_supernet_is_family_mismatch(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """An IPv4 candidate under the IPv6 supernet is rejected with PARENT_SUPERNET_FAMILY_MISMATCH"""
    errors = validate_subnet(
        objects_manager, types_manager, '10.0.0.0/24',
        parent_supernet_id=SUPERNET_ID, subnet_type=IpAddressFamily.IPV4,
    )

    messages = ' '.join(e[ValidationErrorKey.MESSAGE] for e in errors)
    assert 'does not match the address family' in messages
    assert 'of supernet' in messages


# -------------------------------------------------------------------------------------------------------------------- #
#                                          UNASSIGN MODES (IPv6, real Mongo)                                           #
# -------------------------------------------------------------------------------------------------------------------- #
def test_unassign_reference_mode_clears_subnet_ref_against_mongo(
    objects_manager: ObjectsManager, types_manager: TypesManager,
    database_manager: MongoDatabaseManager, database_name: str, full_access_user,
) -> None:
    """reference mode finds the owner via the real interface query, clears the ref, and keeps the row"""
    objects = database_manager.get_collection(CmdbObject.COLLECTION, database_name)
    objects.insert_one(_interface_carrier(UNASSIGN_REF_OWNER_ID, '2001:db8:0:1::6'))

    try:
        result = unassign_ips_from_subnet(
            objects_manager, types_manager, SUBNET_ID, ['2001:db8:0:1::6'], full_access_user,
            raw_mode='reference',
        )

        assert result[IpamUnassignKey.MODE] == IpamUnassignMode.REFERENCE
        assert result[IpamUnassignKey.UNASSIGNED_COUNT] == 1

        stored = objects.find_one({CmdbObjectKey.PUBLIC_ID: UNASSIGN_REF_OWNER_ID})
        row_data = stored[CmdbObjectKey.MULTI_DATA_SECTIONS][0][CmdbObjectMdsKey.VALUES][0][CmdbObjectMdsRowKey.DATA]
        subnet_entry = next(e for e in row_data if e[CmdbObjectFieldKey.NAME] == InterfaceField.SUBNET)
        assert subnet_entry[CmdbObjectFieldKey.VALUE] is None  # ref cleared, the row (IP) is kept
    finally:
        objects.delete_one({CmdbObjectKey.PUBLIC_ID: UNASSIGN_REF_OWNER_ID})


def test_unassign_row_mode_deletes_the_interface_row_against_mongo(
    objects_manager: ObjectsManager, types_manager: TypesManager,
    database_manager: MongoDatabaseManager, database_name: str, full_access_user,
) -> None:
    """row mode deletes the whole matching dg-ipam-interface row from the owner, persisted to Mongo"""
    objects = database_manager.get_collection(CmdbObject.COLLECTION, database_name)
    objects.insert_one(_interface_carrier(UNASSIGN_ROW_OWNER_ID, '2001:db8:0:1::7'))

    try:
        result = unassign_ips_from_subnet(
            objects_manager, types_manager, SUBNET_ID, ['2001:db8:0:1::7'], full_access_user,
            raw_mode='row',
        )

        assert result[IpamUnassignKey.MODE] == IpamUnassignMode.ROW
        assert result[IpamUnassignKey.UNASSIGNED_COUNT] == 1

        stored = objects.find_one({CmdbObjectKey.PUBLIC_ID: UNASSIGN_ROW_OWNER_ID})
        # the interface section survives but its single row was removed (owner object kept)
        assert stored[CmdbObjectKey.MULTI_DATA_SECTIONS][0][CmdbObjectMdsKey.VALUES] == []
    finally:
        objects.delete_one({CmdbObjectKey.PUBLIC_ID: UNASSIGN_ROW_OWNER_ID})


# -------------------------------------------------------------------------------------------------------------------- #
#                                       SUBNET IP EXPORT (IPv6, real Mongo)                                            #
# -------------------------------------------------------------------------------------------------------------------- #
def test_build_subnet_ips_xlsx_ipv6_exports_assigned_only_against_mongo(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The IPv6 subnet export runs the real assigned-rows query and emits only the assigned IP row"""
    content: bytes = build_subnet_ips_xlsx(objects_manager, types_manager, SUBNET_ID)

    sheet = load_workbook(BytesIO(content)).active
    rows: list[tuple[Any, ...]] = list(sheet.iter_rows(values_only=True))

    assert sheet.title == IpamSubnetIpsExport.SHEET_TITLE
    assert rows[0] == tuple(IpamSubnetIpsExport.HEADERS)
    # IPv6 exports assigned addresses only: the seeded carrier at ::5 is the single data row
    assert len(rows) == 2
    assert rows[1][0] == ASSIGNED_IP_V6
    assert rows[1][2] == IpamRowStatus.ASSIGNED


# -------------------------------------------------------------------------------------------------------------------- #
#                                        CANONICALIZE ON STORE (IPv6)                                                  #
# -------------------------------------------------------------------------------------------------------------------- #
def test_enforce_object_invariants_canonicalizes_ipv6_range_and_persists(
    objects_manager: ObjectsManager, types_manager: TypesManager,
    database_manager: MongoDatabaseManager, database_name: str,
) -> None:
    """enforce_object_invariants normalises a non-canonical IPv6 range, and the saved doc stores it canonical"""
    candidate = _object_doc(CANON_SUBNET_ID, SUBNET_TYPE_ID, [
        _field(SubnetField.NAME, 'canon'),
        _field(SubnetField.TYPE, IpAddressFamily.IPV6),
        _field(SubnetField.NETWORK_RANGE, '2001:DB8:0:0::/48'),
    ])

    errors = enforce_object_invariants(objects_manager, types_manager, candidate)

    assert errors == []
    assert extract_field_value(candidate, SubnetField.NETWORK_RANGE) == '2001:db8::/48'

    objects = database_manager.get_collection(CmdbObject.COLLECTION, database_name)
    objects.insert_one(candidate)
    stored = objects.find_one({'public_id': CANON_SUBNET_ID})

    assert extract_field_value(stored, SubnetField.NETWORK_RANGE) == '2001:db8::/48'
