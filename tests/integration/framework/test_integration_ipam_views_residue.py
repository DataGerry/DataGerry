# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2026 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Integration tests for the remaining IPAM view / export / wiring paths against a real MongoDB

Pins the DB-touching behaviour the unit tests only mock: the per-subnet lazy children fetch
(build_supernet_subnet_children), the invalid-subnets-only overview, the supernet overview's
flat search branch, the subnet IP table's status / sort / type-filter query parameters, the
supernet subnets .xlsx export, resolve_supernet_family and validate_vlan's subnet lookup.
The SpecialType ref_types cross-wiring lives in test_integration_ipam_wiring (own module: it
must not share the DB with another SUBNET / SUPERNET SpecialType seed)
"""
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook

from cmdb.database import MongoDatabaseManager
from cmdb.manager import ObjectsManager, TypesManager
from cmdb.models.object_model import (
    CmdbObject,
    CmdbObjectKey,
    CmdbObjectMdsKey,
    CmdbObjectMdsRowKey,
)
from cmdb.models.type_model import CmdbType
from cmdb.models.special_type_model.special_type_enum import SpecialType
from cmdb.models.special_type_model.ipam_constants import (
    SupernetField,
    SubnetField,
    VlanField,
    InterfaceField,
    IpamSection,
    IpAddressFamily,
    IpamOverviewKey,
    IpamRowStatus,
)
from cmdb.framework.ipam.supernet_overview import (
    build_invalid_subnets_overview,
    build_supernet_overview,
    build_supernet_subnet_children,
    resolve_supernet_family,
)
from cmdb.framework.ipam.subnet_overview import build_subnet_overview
from cmdb.framework.ipam.subnet_export import build_supernet_subnets_xlsx
from cmdb.framework.ipam.vlan_validator import VlanErrorCode, validate_vlan
from cmdb.utils import ValidationErrorKey
from tests.utils.ipam_doc_builders import make_field, make_object_doc, make_type_doc
# -------------------------------------------------------------------------------------------------------------------- #

SUPERNET_TYPE_ID: int = 9810
SUBNET_TYPE_ID: int = 9811
CARRIER_A_TYPE_ID: int = 9812
CARRIER_B_TYPE_ID: int = 9813
VLAN_TYPE_ID: int = 9814

SUPERNET_ID: int = 9820
SUBNET_PARENT_ID: int = 9821    # 10.4.0.0/16
SUBNET_CHILD_ID: int = 9822     # 10.4.4.0/24 - CIDR-child of the parent
SUBNET_INVALID_ID: int = 9823   # 192.168.50.0/24 - outside the supernet
SUBNET_IPS_ID: int = 9824       # 10.5.0.0/24 - carries the assigned IPs
CARRIER_A_ID: int = 9825
CARRIER_B_ID: int = 9826
VLAN_ID: int = 9827
FOREIGN_SUBNET_ID: int = 9899   # never seeded

SUPERNET_RANGE: str = '10.0.0.0/8'
SUBNET_PARENT_RANGE: str = '10.4.0.0/16'
SUBNET_CHILD_RANGE: str = '10.4.4.0/24'
SUBNET_INVALID_RANGE: str = '192.168.50.0/24'
SUBNET_IPS_RANGE: str = '10.5.0.0/24'

IP_OF_CARRIER_A: str = '10.5.0.10'
IP_OF_CARRIER_B: str = '10.5.0.20'
VLAN_NAME: str = 'res-vlan'

TYPE_IDS: list[int] = [SUPERNET_TYPE_ID, SUBNET_TYPE_ID, CARRIER_A_TYPE_ID, CARRIER_B_TYPE_ID, VLAN_TYPE_ID]
OBJECT_IDS: list[int] = [
    SUPERNET_ID, SUBNET_PARENT_ID, SUBNET_CHILD_ID, SUBNET_INVALID_ID,
    SUBNET_IPS_ID, CARRIER_A_ID, CARRIER_B_ID, VLAN_ID,
]


# -------------------------------------------------------------------------------------------------------------------- #
#                                                   FIXTURES                                                           #
# -------------------------------------------------------------------------------------------------------------------- #
def _subnet_doc(public_id: int, name: str, cidr: str) -> dict[str, Any]:
    """Builds a SUBNET object doc assigned to the module's supernet."""
    return make_object_doc(public_id, SUBNET_TYPE_ID, [
        make_field(SubnetField.NAME, name),
        make_field(SubnetField.TYPE, IpAddressFamily.IPV4),
        make_field(SubnetField.PARENT_SUPERNET, SUPERNET_ID),
        make_field(SubnetField.NETWORK_RANGE, cidr),
    ])


def _carrier_doc(public_id: int, type_id: int, ip: str) -> dict[str, Any]:
    """Builds a carrier object with one interface row claiming ``ip`` in the IPs subnet."""
    return make_object_doc(public_id, type_id, [make_field('dg-name', f'res-host-{public_id}')], mds=[{
        CmdbObjectMdsKey.SECTION_ID: IpamSection.INTERFACE,
        CmdbObjectMdsKey.VALUES: [{CmdbObjectMdsRowKey.DATA: [
            make_field(InterfaceField.SUBNET, SUBNET_IPS_ID),
            make_field(InterfaceField.IP, ip),
            make_field(InterfaceField.TYPE, IpAddressFamily.IPV4),
        ]}],
    }])


@pytest.fixture(scope='module', autouse=True)
def _seed_views_topology(database_manager: MongoDatabaseManager, database_name: str):
    """Seeds the supernet with nested / invalid / IP-carrying subnets and two carriers."""
    types = database_manager.get_collection(CmdbType.COLLECTION, database_name)
    objects = database_manager.get_collection(CmdbObject.COLLECTION, database_name)

    types.insert_many([
        make_type_doc(SUPERNET_TYPE_ID, 'it-res-supernet', SpecialType.SUPERNET),
        make_type_doc(SUBNET_TYPE_ID, 'it-res-subnet', SpecialType.SUBNET),
        make_type_doc(CARRIER_A_TYPE_ID, 'it-res-carrier-a', None),
        make_type_doc(CARRIER_B_TYPE_ID, 'it-res-carrier-b', None),
        make_type_doc(VLAN_TYPE_ID, 'it-res-vlan', SpecialType.VLAN),
    ])

    objects.insert_many([
        make_object_doc(SUPERNET_ID, SUPERNET_TYPE_ID, [
            make_field(SupernetField.NAME, 'res-sn'),
            make_field(SupernetField.TYPE, IpAddressFamily.IPV4),
            make_field(SupernetField.NETWORK_RANGE, SUPERNET_RANGE),
        ]),
        _subnet_doc(SUBNET_PARENT_ID, 'res-parent', SUBNET_PARENT_RANGE),
        _subnet_doc(SUBNET_CHILD_ID, 'res-child', SUBNET_CHILD_RANGE),
        _subnet_doc(SUBNET_INVALID_ID, 'res-invalid', SUBNET_INVALID_RANGE),
        _subnet_doc(SUBNET_IPS_ID, 'res-ips', SUBNET_IPS_RANGE),
        _carrier_doc(CARRIER_A_ID, CARRIER_A_TYPE_ID, IP_OF_CARRIER_A),
        _carrier_doc(CARRIER_B_ID, CARRIER_B_TYPE_ID, IP_OF_CARRIER_B),
        make_object_doc(VLAN_ID, VLAN_TYPE_ID, [
            make_field(VlanField.NAME, VLAN_NAME),
            make_field(VlanField.SUBNET_REF, SUBNET_PARENT_ID),
        ]),
    ])

    yield

    types.delete_many({CmdbObjectKey.PUBLIC_ID: {'$in': TYPE_IDS}})
    objects.delete_many({CmdbObjectKey.PUBLIC_ID: {'$in': OBJECT_IDS}})


@pytest.fixture(name='objects_manager')
def fixture_objects_manager(database_manager: MongoDatabaseManager) -> ObjectsManager:
    """Provides an ObjectsManager wired to the test database."""
    return ObjectsManager(database_manager)


@pytest.fixture(name='types_manager')
def fixture_types_manager(database_manager: MongoDatabaseManager) -> TypesManager:
    """Provides a TypesManager wired to the test database."""
    return TypesManager(database_manager)


# -------------------------------------------------------------------------------------------------------------------- #
#                                            SUPERNET VIEW RESIDUE                                                     #
# -------------------------------------------------------------------------------------------------------------------- #
def test_subnet_children_fetch_returns_one_nesting_level(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The lazy children fetch returns the /24 as the /16's direct CIDR-child"""
    payload = build_supernet_subnet_children(objects_manager, types_manager, SUPERNET_ID, SUBNET_PARENT_ID)

    child_ids = [row[CmdbObjectKey.PUBLIC_ID] for row in payload[IpamOverviewKey.ROWS]]
    assert child_ids == [SUBNET_CHILD_ID]


def test_invalid_subnet_overview_lists_the_out_of_range_subnet(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The invalid-only view surfaces exactly the subnet whose CIDR falls outside the supernet"""
    payload = build_invalid_subnets_overview(objects_manager, types_manager, SUPERNET_ID)

    invalid_ids = [row[CmdbObjectKey.PUBLIC_ID] for row in payload[IpamOverviewKey.SUBNETS][IpamOverviewKey.ROWS]]
    assert invalid_ids == [SUBNET_INVALID_ID]
    assert payload[IpamOverviewKey.INVALID_COUNT] == 1


def test_supernet_overview_search_drops_the_tree_shape(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """An active search returns the matching nested subnet as a flat row regardless of depth"""
    payload = build_supernet_overview(objects_manager, types_manager, SUPERNET_ID, search='10.4.4')

    rows = payload[IpamOverviewKey.SUBNETS][IpamOverviewKey.ROWS]
    assert [row[CmdbObjectKey.PUBLIC_ID] for row in rows] == [SUBNET_CHILD_ID]


def test_supernet_overview_attaches_vlans_via_the_real_aggregation(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The VLAN grouping aggregation surfaces the seeded VLAN chip on its subnet row"""
    payload = build_supernet_overview(objects_manager, types_manager, SUPERNET_ID)

    rows = payload[IpamOverviewKey.SUBNETS][IpamOverviewKey.ROWS]
    parent_row = next(r for r in rows if r[CmdbObjectKey.PUBLIC_ID] == SUBNET_PARENT_ID)

    assert parent_row[IpamOverviewKey.VLANS] == [
        {CmdbObjectKey.PUBLIC_ID: VLAN_ID, IpamOverviewKey.NAME: VLAN_NAME},
    ]


def test_resolve_supernet_family_loads_the_real_document(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The export helper resolves the supernet's family from its stored CIDR"""
    assert resolve_supernet_family(objects_manager, types_manager, SUPERNET_ID) == IpAddressFamily.IPV4


# -------------------------------------------------------------------------------------------------------------------- #
#                                        SUBNET IP TABLE QUERY PARAMETERS                                              #
# -------------------------------------------------------------------------------------------------------------------- #
def test_subnet_overview_status_filter_keeps_only_assigned_rows(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """status=assigned reduces the IP table to the two stored interface IPs"""
    payload = build_subnet_overview(
        objects_manager, types_manager, SUBNET_IPS_ID, status=IpamRowStatus.ASSIGNED,
    )

    rows = payload[IpamOverviewKey.IPS][IpamOverviewKey.ROWS]
    assert {row[IpamOverviewKey.IP] for row in rows} == {IP_OF_CARRIER_A, IP_OF_CARRIER_B}
    assert all(row[IpamOverviewKey.STATUS] == IpamRowStatus.ASSIGNED for row in rows)


def test_subnet_overview_sort_descending_by_ip(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """sort=ip&order=-1 puts the higher assigned address first"""
    payload = build_subnet_overview(
        objects_manager, types_manager, SUBNET_IPS_ID,
        status=IpamRowStatus.ASSIGNED, sort=IpamOverviewKey.IP, order='-1',
    )

    rows = payload[IpamOverviewKey.IPS][IpamOverviewKey.ROWS]
    assert [row[IpamOverviewKey.IP] for row in rows] == [IP_OF_CARRIER_B, IP_OF_CARRIER_A]


def test_subnet_overview_type_filter_restricts_to_one_owner_type(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """type=<carrier-a id> keeps only the row whose owner is of that CmdbType"""
    payload = build_subnet_overview(
        objects_manager, types_manager, SUBNET_IPS_ID,
        status=IpamRowStatus.ASSIGNED, type_filter=str(CARRIER_A_TYPE_ID),
    )

    rows = payload[IpamOverviewKey.IPS][IpamOverviewKey.ROWS]
    assert {row[IpamOverviewKey.IP] for row in rows} == {IP_OF_CARRIER_A}


# -------------------------------------------------------------------------------------------------------------------- #
#                                              SUPERNET SUBNETS EXPORT                                                 #
# -------------------------------------------------------------------------------------------------------------------- #
def test_supernet_subnets_xlsx_exports_every_assigned_subnet(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """The export workbook carries one row per assigned subnet with its CIDR"""
    content = build_supernet_subnets_xlsx(objects_manager, types_manager, SUPERNET_ID)

    sheet = load_workbook(BytesIO(content)).active
    cidr_column = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]

    assert set(cidr_column) == {
        SUBNET_PARENT_RANGE, SUBNET_CHILD_RANGE, SUBNET_INVALID_RANGE, SUBNET_IPS_RANGE,
    }


# -------------------------------------------------------------------------------------------------------------------- #
#                                                 VLAN VALIDATOR                                                       #
# -------------------------------------------------------------------------------------------------------------------- #
def test_validate_vlan_accepts_an_existing_subnet(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """A vlan referencing a real SUBNET object passes the lookup"""
    assert not validate_vlan(objects_manager, types_manager, SUBNET_PARENT_ID)


def test_validate_vlan_rejects_a_missing_subnet(
    objects_manager: ObjectsManager, types_manager: TypesManager,
) -> None:
    """A vlan referencing a non-existent subnet id is reported as SUBNET_NOT_FOUND"""
    errors = validate_vlan(objects_manager, types_manager, FOREIGN_SUBNET_ID)

    assert len(errors) == 1
    assert errors[0][ValidationErrorKey.CODE] == VlanErrorCode.SUBNET_NOT_FOUND
