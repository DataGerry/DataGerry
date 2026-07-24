# DataGerry - OpenSource Enterprise CMDB
# Copyright (C) 2026 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Unit tests for cmdb.framework.exporter.format.xlsx_export_format
"""
import json
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from cmdb.framework.exporter.format.xlsx_export_format import XlsxExportFormat, MAX_SHEET_TITLE_LENGTH
from cmdb.errors.exporter import ExporterColumnError
# -------------------------------------------------------------------------------------------------------------------- #


def _obj(
        object_id: int,
        type_id: int = 5,
        type_label: str = 'Server',
        fields=None,
        mds=None,
        sections=None) -> SimpleNamespace:
    """A stand-in RenderResult with the given fields (defaulting to one text field) and optional MDS."""
    return SimpleNamespace(
        fields=fields if fields is not None else [{'name': 'dg-name', 'type': 'text', 'value': 'host-1'}],
        sections=sections or [],
        multi_data_sections=mds or [],
        object_information={'object_id': object_id, 'active': True},
        type_information={'type_id': type_id, 'type_label': type_label},
    )


def _mds_section(section_id: str, field_names: list) -> dict:
    """Builds a rendered type-section dict describing a multi-data-section."""
    return {'type': 'multi-data-section', 'name': section_id, 'label': section_id, 'fields': field_names}


def _mds(section_id: str, entries: list) -> dict:
    """Builds an object's MDS section instance from a list of {field_name: value} entry dicts."""
    return {
        'section_id': section_id,
        'highest_id': len(entries),
        'values': [
            {'multi_data_id': idx, 'data': [{'name': name, 'value': val} for name, val in entry.items()]}
            for idx, entry in enumerate(entries)
        ],
    }


def _workbook(data, *args):
    """Runs the XLSX export and loads the resulting bytes back into an openpyxl workbook."""
    return load_workbook(BytesIO(XlsxExportFormat().export(data, *args)))


class TestXlsxExport:
    """XlsxExportFormat.export serializes rendered objects into an XLSX workbook."""

    def test_native_export(self) -> None:
        """One worksheet named after the type, with the header row and one data row."""
        sheet = _workbook([_obj(10)])['Server']

        assert [sheet.cell(1, col).value for col in (1, 2, 3)] == ['public_id', 'active', 'dg-name']
        assert [sheet.cell(2, col).value for col in (1, 2, 3)] == ['10', 'True', 'host-1']

    def test_empty_export_yields_single_header_sheet(self) -> None:
        """An empty object list yields one valid, visible header-only worksheet (no IndexError / 500)."""
        workbook = _workbook([])

        assert len(workbook.sheetnames) == 1
        sheet = workbook.active
        assert [sheet.cell(1, col).value for col in (1, 2)] == ['public_id', 'active']
        assert sheet.cell(2, 1).value is None  # no data rows

    def test_multitype_export_uses_per_type_columns(self) -> None:
        """B2: each type's worksheet carries its own field columns, not the first type's."""
        server = _obj(10, type_id=5, type_label='Server',
                      fields=[{'name': 'dg-name', 'type': 'text', 'value': 'host-1'}])
        router = _obj(11, type_id=6, type_label='Router',
                      fields=[{'name': 'ip', 'type': 'text', 'value': '10.0.0.1'}])

        # Pass unsorted to also exercise the type_id sort
        workbook = _workbook([router, server])

        assert workbook.sheetnames == ['Server', 'Router']

        server_sheet = workbook['Server']
        assert [server_sheet.cell(1, col).value for col in (1, 2, 3)] == ['public_id', 'active', 'dg-name']
        assert [server_sheet.cell(2, col).value for col in (1, 2, 3)] == ['10', 'True', 'host-1']

        router_sheet = workbook['Router']
        assert [router_sheet.cell(1, col).value for col in (1, 2, 3)] == ['public_id', 'active', 'ip']
        assert [router_sheet.cell(2, col).value for col in (1, 2, 3)] == ['11', 'True', '10.0.0.1']

    def test_multiple_objects_of_same_type_share_one_sheet(self) -> None:
        """Two objects of the same type are written as consecutive rows on one worksheet."""
        workbook = _workbook([_obj(10), _obj(11)])

        assert workbook.sheetnames == ['Server']
        sheet = workbook['Server']
        assert sheet.cell(2, 1).value == '10'
        assert sheet.cell(3, 1).value == '11'

    def test_render_metadata_selects_header_and_columns(self) -> None:
        """A render-view metadata override fixes the header identity columns and the field columns."""
        metadata = json.dumps({'header': ['public_id'], 'columns': ['dg-name']})
        sheet = _workbook([_obj(10)], {'view': 'render', 'metadata': metadata})['Server']

        assert [sheet.cell(1, col).value for col in (1, 2)] == ['public_id', 'dg-name']
        assert sheet.cell(1, 3).value is None
        assert [sheet.cell(2, col).value for col in (1, 2)] == ['10', 'host-1']

    def test_no_mds_columns_when_type_has_no_mds(self) -> None:
        """A type whose objects carry no MDS gets no extra columns."""
        sheet = _workbook([_obj(10)])['Server']

        assert sheet.cell(1, 4).value is None

    def test_declares_xlsx_mime_type(self) -> None:
        """XLSX declares the standard spreadsheet mime type (not the writer's text/<ext> fallback)."""
        assert XlsxExportFormat.MIME_TYPE == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class TestXlsxExportMds:
    """MDS fields become their own columns; entries spread across consecutive rows (like CSV)."""

    def test_mds_fields_become_columns_spread_over_rows(self) -> None:
        """Each MDS field is a column; row 1 = regular + entry 1, next rows the following entries."""
        sections = [_mds_section('nics', ['nic_name', 'mac'])]
        mds = [_mds('nics', [{'nic_name': 'eth0', 'mac': 'm0'}, {'nic_name': 'eth1', 'mac': 'm1'}])]

        sheet = _workbook([_obj(10, mds=mds, sections=sections)])['Server']

        assert [sheet.cell(1, c).value for c in (1, 2, 3, 4, 5)] == \
            ['public_id', 'active', 'dg-name', 'nic_name', 'mac']
        assert [sheet.cell(2, c).value for c in (1, 2, 3, 4, 5)] == ['10', 'True', 'host-1', 'eth0', 'm0']
        # Continuation row: identity + regular columns blank, only the 2nd MDS entry present
        assert sheet.cell(3, 1).value in (None, '')
        assert sheet.cell(3, 3).value in (None, '')
        assert [sheet.cell(3, c).value for c in (4, 5)] == ['eth1', 'm1']

    def test_mds_unequal_section_counts(self) -> None:
        """A section with fewer entries leaves its columns blank on the trailing rows."""
        sections = [_mds_section('nics', ['nic_name']), _mds_section('disks', ['disk_label'])]
        mds = [
            _mds('nics', [{'nic_name': 'a'}, {'nic_name': 'b'}, {'nic_name': 'c'}]),
            _mds('disks', [{'disk_label': 'root'}, {'disk_label': 'data'}]),
        ]

        sheet = _workbook([_obj(10, mds=mds, sections=sections)])['Server']

        assert [sheet.cell(1, c).value for c in (4, 5)] == ['nic_name', 'disk_label']
        assert [sheet.cell(2, c).value for c in (4, 5)] == ['a', 'root']
        assert [sheet.cell(3, c).value for c in (4, 5)] == ['b', 'data']
        # 3rd entry: nics has one, disks does not
        assert sheet.cell(4, 4).value == 'c'
        assert sheet.cell(4, 5).value in (None, '')

    def test_mds_columns_are_per_type(self) -> None:
        """MDS columns appear only on the worksheet of the type that defines them."""
        sections = [_mds_section('nics', ['nic_name'])]
        server = _obj(10, type_id=5, type_label='Server',
                      mds=[_mds('nics', [{'nic_name': 'eth0'}])], sections=sections)
        router = _obj(11, type_id=6, type_label='Router')

        workbook = _workbook([server, router])

        assert workbook['Server'].cell(1, 4).value == 'nic_name'
        assert workbook['Router'].cell(1, 4).value is None

    def test_mds_field_not_duplicated_from_flat_fields(self) -> None:
        """An MDS field also present (default-valued) in the flat fields is emitted once, with MDS values."""
        sections = [_mds_section('nics', ['nic_name'])]
        fields = [{'name': 'dg-name', 'type': 'text', 'value': 'host-1'},
                  {'name': 'nic_name', 'type': 'text', 'value': 'DEFAULT'}]
        mds = [_mds('nics', [{'nic_name': 'eth0'}])]

        sheet = _workbook([_obj(10, mds=mds, sections=sections, fields=fields)])['Server']

        assert [sheet.cell(1, c).value for c in (1, 2, 3, 4)] == ['public_id', 'active', 'dg-name', 'nic_name']
        assert sheet.cell(1, 5).value is None  # nic_name appears once
        assert sheet.cell(2, 4).value == 'eth0'

    def test_duplicate_column_name_raises(self) -> None:
        """Two fields resolving to the same column name refuse the export."""
        sections = [_mds_section('s1', ['dup']), _mds_section('s2', ['dup'])]

        with pytest.raises(ExporterColumnError):
            XlsxExportFormat().export([_obj(10, sections=sections)])


class TestXlsxExportHumanReadable:
    """The human_readable flag relabels headers and resolves reference / location values."""

    def test_headers_relabelled_and_values_resolved(self, human_readable_object) -> None:
        """Headers become labels (incl. identity); ref -> summary line; location -> tree name."""
        options = {'human_readable': 'true', 'location_names': {42: 'Berlin/Room-1'}}
        sheet = _workbook([human_readable_object], options)['Server']

        assert [sheet.cell(1, c).value for c in (1, 2, 3, 4, 5)] == \
            ['Public ID', 'Active', 'Hostname', 'Owner', 'Location']
        assert [sheet.cell(2, c).value for c in (1, 2, 3, 4, 5)] == \
            ['10', 'True', 'host-1', 'User #3 | alice', 'Berlin/Room-1']


class TestNormalizeSheetTitle:
    """XlsxExportFormat._normalize_sheet_title sanitizes a type label for use as a sheet title."""
    # pylint: disable=protected-access

    def test_replaces_invalid_characters(self) -> None:
        """Excel-invalid characters are replaced with underscores."""
        assert XlsxExportFormat._normalize_sheet_title('a/b:c*d?e[f]g\\h') == 'a_b_c_d_e_f_g_h'

    def test_truncates_to_excel_limit(self) -> None:
        """A title longer than Excel's 31-character limit is truncated."""
        assert XlsxExportFormat._normalize_sheet_title('x' * 50) == 'x' * MAX_SHEET_TITLE_LENGTH
