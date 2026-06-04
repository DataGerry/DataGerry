# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2026 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Unit tests for cmdb.framework.ipam.subnet_export

Covers the IP-range cell formatting, the per-row mapping to export cells, and the full workbook
build: the assigned subnet rows are stubbed (load_assigned_subnet_rows is patched) and the produced
.xlsx bytes are read back with openpyxl to assert the sheet title, header row and data rows.
"""
from typing import Any
from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from cmdb.models.special_type_model.ipam_constants import (
    IpamOverviewKey,
    IpamExport,
    IpamSubnetIpsExport,
    IpamRowStatus,
    IpAddressFamily,
)
from cmdb.framework.ipam.subnet_export import (
    _format_ip_range,
    _subnet_export_row,
    _subnet_ip_export_row,
    build_supernet_subnets_xlsx,
    build_subnet_ips_xlsx,
)
# -------------------------------------------------------------------------------------------------------------------- #

MODULE: str = 'cmdb.framework.ipam.subnet_export'

ROW_A: dict[str, Any] = {
    IpamOverviewKey.CIDR: '10.0.0.0/24',
    IpamOverviewKey.IP_RANGE: {IpamOverviewKey.FIRST: '10.0.0.0', IpamOverviewKey.LAST: '10.0.0.255'},
    IpamOverviewKey.USED_IPS: 3,
    IpamOverviewKey.FREE_IPS: 253,
    IpamOverviewKey.USAGE_PERCENT: 1.17,
}
ROW_DEGENERATE: dict[str, Any] = {
    IpamOverviewKey.CIDR: 'not-a-cidr',
    IpamOverviewKey.IP_RANGE: None,
    IpamOverviewKey.USED_IPS: 0,
    IpamOverviewKey.FREE_IPS: 0,
    IpamOverviewKey.USAGE_PERCENT: 0.0,
}

# -------------------------------------------------------------------------------------------------------------------- #
#                                              _format_ip_range                                                      #
# -------------------------------------------------------------------------------------------------------------------- #

def test_format_ip_range_joins_first_and_last() -> None:
    """A populated range renders as 'first - last'"""
    rendered: str = _format_ip_range({IpamOverviewKey.FIRST: '10.0.0.0', IpamOverviewKey.LAST: '10.0.0.255'})

    assert rendered == '10.0.0.0 - 10.0.0.255'


def test_format_ip_range_returns_empty_for_missing_range() -> None:
    """A None / empty range renders as an empty string"""
    assert _format_ip_range(None) == ''
    assert _format_ip_range({}) == ''

# -------------------------------------------------------------------------------------------------------------------- #
#                                             _subnet_export_row                                                     #
# -------------------------------------------------------------------------------------------------------------------- #

def test_subnet_export_row_includes_usage_for_ipv4() -> None:
    """With include_usage=True the row carries the trailing usage-percent cell (IPv4 export)"""
    assert _subnet_export_row(ROW_A, include_usage=True) == ['10.0.0.0/24', '10.0.0.0 - 10.0.0.255', 3, 253, 1.17]


def test_subnet_export_row_omits_usage_for_ipv6() -> None:
    """With include_usage=False the row stops at free_ips (IPv6 export drops the usage column)"""
    assert _subnet_export_row(ROW_A, include_usage=False) == ['10.0.0.0/24', '10.0.0.0 - 10.0.0.255', 3, 253]

# -------------------------------------------------------------------------------------------------------------------- #
#                                        build_supernet_subnets_xlsx                                                 #
# -------------------------------------------------------------------------------------------------------------------- #

def test_build_supernet_subnets_xlsx_ipv4_includes_usage_column() -> None:
    """An IPv4 supernet's sheet carries the trailing 'Usage (%)' header and per-row usage cell"""
    with patch(f'{MODULE}.resolve_supernet_family', return_value=IpAddressFamily.IPV4), \
         patch(f'{MODULE}.load_assigned_subnet_rows', return_value=[ROW_A, ROW_DEGENERATE]):
        content: bytes = build_supernet_subnets_xlsx(MagicMock(), MagicMock(), 42)

    sheet = load_workbook(BytesIO(content)).active
    assert sheet.title == IpamExport.SHEET_TITLE

    rows: list[tuple[Any, ...]] = list(sheet.iter_rows(values_only=True))
    assert rows[0] == tuple(IpamExport.HEADERS + [IpamExport.USAGE_HEADER])
    assert rows[1] == ('10.0.0.0/24', '10.0.0.0 - 10.0.0.255', 3, 253, 1.17)
    assert rows[2] == ('not-a-cidr', None, 0, 0, 0.0)


def test_build_supernet_subnets_xlsx_ipv6_omits_usage_column() -> None:
    """An IPv6 supernet's sheet has the base headers only and no per-row usage cell"""
    row_v6: dict[str, Any] = {
        IpamOverviewKey.CIDR: '2001:db8:1::/64',
        IpamOverviewKey.IP_RANGE: {IpamOverviewKey.FIRST: '2001:db8:1::', IpamOverviewKey.LAST: '2001:db8:1::ffff'},
        IpamOverviewKey.USED_IPS: 1,
        IpamOverviewKey.FREE_IPS: 18446744073709551615,
        IpamOverviewKey.USAGE_PERCENT: None,
    }

    with patch(f'{MODULE}.resolve_supernet_family', return_value=IpAddressFamily.IPV6), \
         patch(f'{MODULE}.load_assigned_subnet_rows', return_value=[row_v6]):
        content: bytes = build_supernet_subnets_xlsx(MagicMock(), MagicMock(), 42)

    rows: list[tuple[Any, ...]] = list(load_workbook(BytesIO(content)).active.iter_rows(values_only=True))
    # No 'Usage (%)' column for IPv6: header + data row are both 4 cells (the huge free_ips count
    # loses precision through Excel's float64, so only the bounded cells are asserted exactly)
    assert rows[0] == tuple(IpamExport.HEADERS)
    assert len(rows[1]) == 4
    assert rows[1][:3] == ('2001:db8:1::/64', '2001:db8:1:: - 2001:db8:1::ffff', 1)


def test_build_supernet_subnets_xlsx_emits_header_only_when_no_subnets() -> None:
    """With no assigned subnets the workbook still carries just the (family-appropriate) header row"""
    with patch(f'{MODULE}.resolve_supernet_family', return_value=IpAddressFamily.IPV4), \
         patch(f'{MODULE}.load_assigned_subnet_rows', return_value=[]):
        content: bytes = build_supernet_subnets_xlsx(MagicMock(), MagicMock(), 42)

    rows: list[tuple[Any, ...]] = list(load_workbook(BytesIO(content)).active.iter_rows(values_only=True))

    assert rows == [tuple(IpamExport.HEADERS + [IpamExport.USAGE_HEADER])]


# -------------------------------------------------------------------------------------------------------------------- #
#                                            _subnet_ip_export_row                                                    #
# -------------------------------------------------------------------------------------------------------------------- #

IP_ROW_ASSIGNED: dict[str, Any] = {
    IpamOverviewKey.IP: '10.0.0.5',
    IpamOverviewKey.STATUS: IpamRowStatus.ASSIGNED,
    IpamOverviewKey.TYPE_INFO: {IpamOverviewKey.LABEL: 'Server', IpamOverviewKey.CI_EXPLORER_COLOR: '#fff'},
    IpamOverviewKey.ASSIGNED_TO: {IpamOverviewKey.SUMMARY_LINE: 'Server: web01'},
    IpamOverviewKey.MAC_ADDRESS: 'aa:bb:cc:dd:ee:ff',
}
IP_ROW_FREE: dict[str, Any] = {
    IpamOverviewKey.IP: '10.0.0.6',
    IpamOverviewKey.STATUS: IpamRowStatus.FREE,
    IpamOverviewKey.TYPE_INFO: None,
    IpamOverviewKey.ASSIGNED_TO: None,
    IpamOverviewKey.MAC_ADDRESS: None,
}


def test_subnet_ip_export_row_maps_assigned_row_to_human_readable_cells() -> None:
    """An assigned row carries the type label, status, owner summary line and MAC"""
    assert _subnet_ip_export_row(IP_ROW_ASSIGNED) == [
        '10.0.0.5', 'Server', IpamRowStatus.ASSIGNED, 'Server: web01', 'aa:bb:cc:dd:ee:ff',
    ]


def test_subnet_ip_export_row_blanks_type_owner_and_mac_for_free_row() -> None:
    """A free row leaves the type, assigned-to and MAC cells blank (not None)"""
    assert _subnet_ip_export_row(IP_ROW_FREE) == ['10.0.0.6', '', IpamRowStatus.FREE, '', '']


# -------------------------------------------------------------------------------------------------------------------- #
#                                            build_subnet_ips_xlsx                                                    #
# -------------------------------------------------------------------------------------------------------------------- #

def test_build_subnet_ips_xlsx_writes_headers_and_rows() -> None:
    """The workbook carries the IP-export sheet title, header row and one data row per IP"""
    with patch(f'{MODULE}.build_subnet_ip_export_rows', return_value=[IP_ROW_ASSIGNED, IP_ROW_FREE]):
        content: bytes = build_subnet_ips_xlsx(MagicMock(), MagicMock(), 7)

    sheet = load_workbook(BytesIO(content)).active
    assert sheet.title == IpamSubnetIpsExport.SHEET_TITLE

    rows: list[tuple[Any, ...]] = list(sheet.iter_rows(values_only=True))
    assert rows[0] == tuple(IpamSubnetIpsExport.HEADERS)
    assert rows[1] == ('10.0.0.5', 'Server', IpamRowStatus.ASSIGNED, 'Server: web01', 'aa:bb:cc:dd:ee:ff')
    # the free row's blank cells round-trip back through openpyxl as empty (None) cells
    assert rows[2] == ('10.0.0.6', None, IpamRowStatus.FREE, None, None)


def test_build_subnet_ips_xlsx_emits_header_only_when_no_rows() -> None:
    """With no exportable IPs the workbook still carries just the header row"""
    with patch(f'{MODULE}.build_subnet_ip_export_rows', return_value=[]):
        content: bytes = build_subnet_ips_xlsx(MagicMock(), MagicMock(), 7)

    rows: list[tuple[Any, ...]] = list(load_workbook(BytesIO(content)).active.iter_rows(values_only=True))

    assert rows == [tuple(IpamSubnetIpsExport.HEADERS)]
