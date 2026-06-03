# DATAGERRY - OpenSource Enterprise CMDB
# Copyright (C) 2026 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Unit tests for cmdb.interface.rest_api.routes.ipam_routes.ipam_supernet_routes

Covers the route-glue responsibilities of get_supernet_overview: reading page / page_size /
search from the query string under their IpamOverviewKey constants, applying the
IpamSearch.MAX_QUERY_LENGTH truncation cap, and forwarding the values to the framework-layer
orchestrator. Substantive behavior (top-level vs. flat list, minimum query length, KPI
computation) belongs to build_supernet_overview and is covered in the framework-layer tests;
this module only exercises the transport boundary

The route function carries auth decorators that abort outside a real session, so each test
unwraps the decorator chain via __wrapped__ and calls the bare handler inside a Flask
test_request_context. build_supernet_overview and ManagerProvider.get_manager are patched at
the route module path so no DB or business logic runs
"""
from typing import Any, Callable
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from flask import Flask
from openpyxl import load_workbook

from cmdb.models.special_type_model.ipam_constants import (
    IpamPagination,
    IpamSearch,
    IpamOverviewKey,
    IpamExport,
    IpAddressFamily,
)
from cmdb.interface.rest_api.routes.ipam_routes.ipam_supernet_routes import (
    get_supernet_overview,
    export_supernet_subnets,
)
# -------------------------------------------------------------------------------------------------------------------- #

ROUTE_PATH: str = 'cmdb.interface.rest_api.routes.ipam_routes.ipam_supernet_routes'
SUPERNET_PUBLIC_ID: int = 42


def _unwrap(func: Callable[..., Any]) -> Callable[..., Any]:
    """Strips the @verify_api_access / @insert_request_user decorators off a route function."""
    inner = func

    while hasattr(inner, '__wrapped__'):
        inner = inner.__wrapped__

    return inner


@pytest.fixture(name='bare_get_supernet_overview')
def fixture_bare_get_supernet_overview() -> Callable[..., Any]:
    """Returns the undecorated get_supernet_overview handler, callable inside a request context."""
    return _unwrap(get_supernet_overview)


@pytest.fixture(name='flask_app')
def fixture_flask_app() -> Flask:
    """Returns a minimal Flask app to host the test_request_context calls."""
    return Flask(__name__)


@pytest.fixture(name='patched_orchestrator')
def fixture_patched_orchestrator() -> Any:
    """
    Patches build_supernet_overview and ManagerProvider.get_manager at the route module path

    The orchestrator returns a sentinel payload so the route's DefaultResponse wrapper has
    something to serialise; tests assert against the captured kwargs on the orchestrator mock
    """
    payload: dict[str, Any] = {'sentinel': True}

    with patch(f'{ROUTE_PATH}.build_supernet_overview', return_value=payload) as mock_build, \
         patch(f'{ROUTE_PATH}.ManagerProvider.get_manager', return_value=MagicMock()):
        yield mock_build


# -------------------------------------------------------------------------------------------------------------------- #
#                                            search query-param forwarding                                             #
# -------------------------------------------------------------------------------------------------------------------- #
def test_route_forwards_search_query_param_to_orchestrator(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """A 'search=10.0' query param is read under IpamOverviewKey.SEARCH and forwarded verbatim"""
    with flask_app.test_request_context('/?search=10.0'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert patched_orchestrator.call_args.kwargs['search'] == '10.0'


def test_route_forwards_empty_search_when_query_param_missing(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """Missing 'search' query param → empty string forwarded so the framework restores top-level view"""
    with flask_app.test_request_context('/'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert patched_orchestrator.call_args.kwargs['search'] == ''


def test_route_truncates_search_at_max_query_length(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """A search longer than IpamSearch.MAX_QUERY_LENGTH is truncated at the route boundary"""
    oversized: str = 'x' * (IpamSearch.MAX_QUERY_LENGTH + 50)

    with flask_app.test_request_context(f'/?search={oversized}'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    forwarded: str = patched_orchestrator.call_args.kwargs['search']
    assert len(forwarded) == IpamSearch.MAX_QUERY_LENGTH
    assert forwarded == 'x' * IpamSearch.MAX_QUERY_LENGTH


def test_route_passes_through_search_at_or_below_max_query_length(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """A search exactly at MAX_QUERY_LENGTH is forwarded unchanged - truncation is open-ended only"""
    at_limit: str = 'a' * IpamSearch.MAX_QUERY_LENGTH

    with flask_app.test_request_context(f'/?search={at_limit}'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert patched_orchestrator.call_args.kwargs['search'] == at_limit


# -------------------------------------------------------------------------------------------------------------------- #
#                                          page / page_size forwarding                                                 #
# -------------------------------------------------------------------------------------------------------------------- #
def test_route_forwards_default_page_and_page_size_when_not_provided(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """Missing page / page_size query params → 1 and IpamPagination.DEFAULT_PAGE_SIZE forwarded"""
    with flask_app.test_request_context('/'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert patched_orchestrator.call_args.kwargs['page'] == 1
    assert patched_orchestrator.call_args.kwargs['page_size'] == IpamPagination.DEFAULT_PAGE_SIZE


def test_route_reads_page_and_page_size_from_request_args_under_enum_keys(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """page / page_size are read under their IpamOverviewKey constants and forwarded as integers"""
    assert IpamOverviewKey.PAGE == 'page'
    assert IpamOverviewKey.PAGE_SIZE == 'page_size'

    with flask_app.test_request_context('/?page=3&page_size=25'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert patched_orchestrator.call_args.kwargs['page'] == 3
    assert patched_orchestrator.call_args.kwargs['page_size'] == 25


def test_route_forwards_public_id_argument_to_orchestrator(
    bare_get_supernet_overview: Callable[..., Any],
    flask_app: Flask,
    patched_orchestrator: Any,
) -> None:
    """The path's public_id flows through to build_supernet_overview as the third positional arg"""
    with flask_app.test_request_context('/'):
        bare_get_supernet_overview(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert patched_orchestrator.call_args.args[2] == SUPERNET_PUBLIC_ID


# -------------------------------------------------------------------------------------------------------------------- #
#                                          export_supernet_subnets                                                     #
# -------------------------------------------------------------------------------------------------------------------- #
@pytest.fixture(name='bare_export_supernet_subnets')
def fixture_bare_export_supernet_subnets() -> Callable[..., Any]:
    """Returns the undecorated export_supernet_subnets handler."""
    return _unwrap(export_supernet_subnets)


def test_export_route_returns_xlsx_attachment_download(
    bare_export_supernet_subnets: Callable[..., Any],
    flask_app: Flask,
) -> None:
    """The route returns the builder's bytes as an .xlsx attachment with the OpenXML mimetype"""
    content: bytes = b'fake-xlsx-bytes'

    with patch(f'{ROUTE_PATH}.build_supernet_subnets_xlsx', return_value=content) as mock_build, \
         patch(f'{ROUTE_PATH}.ManagerProvider.get_manager', return_value=MagicMock()), \
         flask_app.test_request_context('/'):
        response = bare_export_supernet_subnets(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    assert response.mimetype == IpamExport.MIMETYPE
    assert response.data == content

    disposition: str = response.headers['Content-Disposition']
    assert disposition.startswith('attachment; filename=')
    assert f'supernet_{SUPERNET_PUBLIC_ID}_subnets_' in disposition
    assert disposition.endswith('.xlsx')

    mock_build.assert_called_once()


def test_export_route_body_parses_as_a_valid_xlsx_workbook(
    bare_export_supernet_subnets: Callable[..., Any],
    flask_app: Flask,
) -> None:
    """End-to-end: with the real builder, the route's body is a parseable .xlsx with the expected rows"""
    rows: list[dict[str, Any]] = [{
        IpamOverviewKey.CIDR: '10.0.0.0/24',
        IpamOverviewKey.IP_RANGE: {IpamOverviewKey.FIRST: '10.0.0.0', IpamOverviewKey.LAST: '10.0.0.255'},
        IpamOverviewKey.USED_IPS: 3,
        IpamOverviewKey.FREE_IPS: 253,
        IpamOverviewKey.USAGE_PERCENT: 1.17,
    }]

    # Patch the data source + the family resolver (IPv4) so the real build runs and emits real bytes
    with patch('cmdb.framework.ipam.subnet_export.load_assigned_subnet_rows', return_value=rows), \
         patch('cmdb.framework.ipam.subnet_export.resolve_supernet_family', return_value=IpAddressFamily.IPV4), \
         patch(f'{ROUTE_PATH}.ManagerProvider.get_manager', return_value=MagicMock()), \
         flask_app.test_request_context('/'):
        response = bare_export_supernet_subnets(public_id=SUPERNET_PUBLIC_ID, request_user=MagicMock())

    sheet = load_workbook(BytesIO(response.data)).active
    sheet_rows: list[tuple[Any, ...]] = list(sheet.iter_rows(values_only=True))

    assert sheet.title == IpamExport.SHEET_TITLE
    # IPv4 supernet export keeps the trailing 'Usage (%)' column
    assert sheet_rows[0] == tuple(IpamExport.HEADERS + [IpamExport.USAGE_HEADER])
    assert sheet_rows[1] == ('10.0.0.0/24', '10.0.0.0 - 10.0.0.255', 3, 253, 1.17)
