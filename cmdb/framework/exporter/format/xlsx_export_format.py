# DataGerry - OpenSource Enterprise CMDB
# Copyright (C) 2026 becon GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
Implementation of XlsxExportFormat
"""
from logging import Logger, getLogger
from io import BytesIO
from collections import namedtuple
import re
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cmdb.models.object_model.cmdb_object_key_enum import CmdbObjectKey
from cmdb.models.type_model.field_key_enum import FieldKey
from cmdb.framework.exporter.format.base_exporter_format import (
    BaseExporterFormat,
    TYPE_INFO_ID_KEY,
    TYPE_INFO_LABEL_KEY,
)
from cmdb.framework.exporter.config.exporter_config_type_enum import ExporterConfigType
from cmdb.framework.exporter.exporter_constants import ExporterMetadataKey, ExporterOptionKey
from cmdb.framework.rendering.render_result import RenderResult
# -------------------------------------------------------------------------------------------------------------------- #

LOGGER: Logger = getLogger(__name__)

# Default identity columns emitted for each object (public_id + active flag)
DEFAULT_HEADER: list[str] = [CmdbObjectKey.PUBLIC_ID.value, CmdbObjectKey.ACTIVE.value]

# Characters not allowed in an Excel sheet title, plus Excel's hard 31-character title limit
INVALID_SHEET_TITLE_CHARS: str = r'[\\*?:/\[\]]'
MAX_SHEET_TITLE_LENGTH: int = 31

# Per-export settings shared by every worksheet (bundled to keep method signatures small)
_SheetSettings = namedtuple(
    '_SheetSettings', ['header', 'metadata_columns', 'view', 'human_readable', 'location_names']
)

# -------------------------------------------------------------------------------------------------------------------- #
#                                               XlsxExportFormat - CLASS                                               #
# -------------------------------------------------------------------------------------------------------------------- #
class XlsxExportFormat(BaseExporterFormat):
    """
    The XLSX export format class for exporting data to Excel (.xlsx) files

    Objects are grouped onto one worksheet per type (sorted by type id); each worksheet carries the
    identity header columns, that type's regular field columns, and one column per multi-data-section
    (MDS) field. MDS entries are spread over consecutive rows the same way the CSV export does (see
    `BaseExporterFormat.build_object_rows`).

    Extends: BaseExporterFormat
    """
    FILE_EXTENSION = "xlsx"
    MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    LABEL = "XLSX"
    MULTITYPE_SUPPORT = True
    ICON = "file-excel"
    DESCRIPTION = "Export as XLSX"
    ACTIVE = True


    def export(self, data: list[RenderResult], *args) -> bytes:
        """
        Exports a list of RenderResult objects as an XLSX file

        Args:
            data (list[RenderResult]): The objects to be exported
            *args: Optional export parameters dict (`view`, `metadata`)

        Returns:
            bytes: The content of the XLSX file as a byte string
        """
        workbook: Workbook = self.create_xls_object(data, args)

        buffer = BytesIO()
        workbook.save(buffer)

        return buffer.getvalue()


    def create_xls_object(self, data: list[RenderResult], args: tuple) -> Workbook:
        """
        Creates an XLSX workbook with the provided data

        Objects are sorted by type id and written onto one worksheet per type. In the NATIVE view each
        worksheet's field columns are the field names of that type (so a multi-type export keeps every
        type's own fields); a render-view `metadata` override instead fixes the header/columns across all
        worksheets. An empty object list still yields one valid, visible header-only worksheet.

        Args:
            data (list[RenderResult]): The objects to be exported
            args (tuple): The positional export args; `args[0]` (if present) is the options dict

        Returns:
            Workbook: The created XLSX workbook
        """
        workbook = Workbook()
        workbook.remove(workbook.active)  # drop the empty default sheet openpyxl creates

        settings = self._resolve_settings(args)

        sorted_data = sorted(data, key=lambda obj: obj.type_information[TYPE_INFO_ID_KEY])
        self._write_type_worksheets(workbook, sorted_data, settings)

        # openpyxl refuses to save a workbook with no visible sheet, so an empty export gets a header-only one
        if not workbook.sheetnames:
            self._write_row(workbook.create_sheet(), 1, list(settings.header))

        return workbook


    def _write_type_worksheets(
            self,
            workbook: Workbook,
            sorted_data: list[RenderResult],
            settings: '_SheetSettings') -> None:
        """
        Writes one worksheet per type into the workbook

        `sorted_data` must be sorted by type id; each contiguous run of one type becomes a worksheet.

        Args:
            workbook (Workbook): The workbook to write worksheets into
            sorted_data (list[RenderResult]): The objects to export, sorted by type id
            settings (_SheetSettings): The shared per-export settings
        """
        for type_objects in self._group_by_type(sorted_data):
            self._write_type_sheet(workbook, type_objects, settings)


    @staticmethod
    def _group_by_type(sorted_data: list[RenderResult]) -> list[list[RenderResult]]:
        """
        Groups type-id-sorted objects into contiguous per-type blocks

        Args:
            sorted_data (list[RenderResult]): The objects, sorted by type id

        Returns:
            list[list[RenderResult]]: One list of objects per type, in type-id order
        """
        groups: list[list[RenderResult]] = []
        current_type_id = None

        for obj in sorted_data:
            type_id = obj.type_information[TYPE_INFO_ID_KEY]

            if current_type_id != type_id or not groups:
                current_type_id = type_id
                groups.append([])

            groups[-1].append(obj)

        return groups


    def _write_type_sheet(
            self,
            workbook: Workbook,
            type_objects: list[RenderResult],
            settings: '_SheetSettings') -> None:
        """
        Writes one type's objects onto a dedicated worksheet

        The worksheet header is the identity columns, the type's regular field columns and one column per
        MDS field; each object then contributes one or more rows (its MDS entries spread over rows). In a
        HUMAN_READABLE export the header cells are relabelled (field labels) as the last step.

        Args:
            workbook (Workbook): The workbook to add the worksheet to
            type_objects (list[RenderResult]): The objects of a single type
            settings (_SheetSettings): The shared per-export settings

        Raises:
            ExporterColumnError: If two fields resolve to the same column name
        """
        first = type_objects[0]
        regular_columns, mds_layout, mds_columns = self._resolve_columns(first, settings.metadata_columns)

        titles: list[str] = [*settings.header, *regular_columns, *mds_columns]
        # The duplicate guard runs on the unique field NAMES, before any human-readable relabel
        self.assert_unique_columns(titles)

        sheet = workbook.create_sheet(title=self._normalize_sheet_title(first.type_information[TYPE_INFO_LABEL_KEY]))
        self._write_row(sheet, 1, self.relabel_header(titles, type_objects) if settings.human_readable else titles)

        row_index = 2  # data rows start below the header
        for obj in type_objects:
            for row in self.build_object_rows(obj, settings.header, regular_columns, mds_layout,
                                              settings.view, settings.human_readable, settings.location_names):
                self._write_row(sheet, row_index, row)
                row_index += 1


    def _resolve_columns(
            self,
            first: RenderResult,
            metadata_columns: list[str] | None) -> tuple[list[str], list[tuple[str, list[str]]], list[str]]:
        """
        Resolves a worksheet's regular field columns, MDS layout and MDS columns for a type

        Args:
            first (RenderResult): The first object of the type (its sections/fields define the columns)
            metadata_columns (list[str] | None): Fixed field columns, or None for the type's own fields

        Returns:
            tuple[list[str], list[tuple[str, list[str]]], list[str]]:
                - regular_columns: the regular (non-MDS) field columns, in output order
                - mds_layout: the `(section_id, field_names)` layout of the type
                - mds_columns: the flattened MDS field columns, in layout order
        """
        mds_layout = self.extract_mds_layout(first.sections)
        mds_columns: list[str] = [name for _, field_names in mds_layout for name in field_names]
        mds_column_set: set[str] = set(mds_columns)

        base_columns = metadata_columns if metadata_columns is not None else self._field_names(first)
        # MDS fields also appear (default-valued) in the flat field list; emit each once as its MDS column
        regular_columns = [name for name in base_columns if name not in mds_column_set]

        return regular_columns, mds_layout, mds_columns


    def _resolve_settings(self, args: tuple) -> '_SheetSettings':
        """
        Resolves the shared per-export settings from the export args

        A render-view `metadata` override supplies the header and the fixed columns used for every
        worksheet; without such an override the export is forced to the NATIVE view and the columns are
        derived per worksheet from each type's fields (`metadata_columns` is None). The HUMAN_READABLE
        flag and the resolved location-name map are read from the options too.

        Args:
            args (tuple): The positional export args; `args[0]` (if present) is the options dict

        Returns:
            _SheetSettings: The bundled (header, metadata_columns, view, human_readable, location_names)
        """
        view, metadata = BaseExporterFormat.resolve_export_view(args)
        options = args[0] if args else {}

        header: list[str] = list(DEFAULT_HEADER)
        metadata_columns: list[str] | None = None

        if metadata:
            header = metadata.get(ExporterMetadataKey.HEADER.value, header)
            metadata_columns = metadata.get(ExporterMetadataKey.COLUMNS.value, [])
        else:
            # XLSX renders in the render view only when metadata explicitly selects the columns
            view = ExporterConfigType.NATIVE.value

        return _SheetSettings(
            header=header,
            metadata_columns=metadata_columns,
            view=view,
            human_readable=self.is_human_readable(options),
            location_names=options.get(ExporterOptionKey.LOCATION_NAMES.value) or {},
        )


    @staticmethod
    def _field_names(obj: RenderResult) -> list[str]:
        """
        Returns the field names of a single object in definition order

        Args:
            obj (RenderResult): The object to read the field names from

        Returns:
            list[str]: The object's field names
        """
        return [field[FieldKey.NAME.value] for field in obj.fields]


    @staticmethod
    def _write_row(sheet: Worksheet, row_index: int, values: list[str]) -> None:
        """
        Writes a single row of pre-stringified cell values into a worksheet

        Args:
            sheet (Worksheet): The worksheet to write into
            row_index (int): The 1-based row number to write on
            values (list[str]): The ordered cell values for the row
        """
        for col_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col_index).value = value


    @staticmethod
    def _normalize_sheet_title(input_data: str) -> str:
        """
        Normalizes a sheet title by replacing invalid characters and enforcing Excel's length limit

        Args:
            input_data (str): The raw sheet title

        Returns:
            str: The normalized sheet title (invalid characters replaced, truncated to 31 characters)
        """
        normalized = re.sub(INVALID_SHEET_TITLE_CHARS, '_', input_data)

        return normalized[:MAX_SHEET_TITLE_LENGTH]
